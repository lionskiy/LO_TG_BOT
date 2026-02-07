"""
Excel import for HR employees: parse sheets ДДЖ and Инфоком, validate, merge by personal_number.
SPEC_HR_SERVICE sections 3-4. Supports .xlsx (openpyxl) and .xls (xlrd).
"""
import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from datetime import date

from api.employees_repository import (
    create_employee,
    employee_exists_by_personal_number,
    _parse_date,
)

logger = logging.getLogger(__name__)

SHEET_DDJ = "ДДЖ"
SHEET_INFOKOM = "Инфоком"

# Column headers (exact or normalized) -> field name
# ДДЖ
DDJ_COLUMNS = {
    "табельный номер": "personal_number",
    "фои": "full_name",
    "фио": "full_name",
    "должность": "position",
    "текст основное мвз": "mvz",
    "дата приема": "hire_date",
    "почта": "email",
    "руководитель": "supervisor",
}
# Инфоком: "Табельный №" first occurrence = personal_number, "Табельный номер" = full_name (FIO!)
# We identify by header text and order for "Табельный №" (first vs second)
INFOKOM_POSITION_COL = "полное наименование штатной до"
INFOKOM_PERSONAL_FIRST = "табельный №"  # first occurrence by column order
INFOKOM_FULL_NAME_COL = "табельный номер"  # on Infokom this column contains FIO
INFOKOM_HIRE = "дата приема"
INFOKOM_SUPERVISOR = "табельный номер начальника"
INFOKOM_SUPERVISOR_ALT = "фио руководителя"
INFOKOM_MVZ = "мвз (название)"
INFOKOM_EMAIL = "почта"


def _normalize_header(s: str) -> str:
    """Lowercase, strip, collapse spaces."""
    if s is None or not isinstance(s, str):
        return ""
    return " ".join(re.split(r"\s+", s.lower().strip()))


def _cell_value(row: List, idx: int) -> str:
    """Safe cell value as string."""
    if idx is None or idx < 0 or idx >= len(row):
        return ""
    v = row[idx]
    if v is None:
        return ""
    if isinstance(v, (datetime,)):
        return v.strftime("%Y-%m-%d") if hasattr(v, "strftime") else str(v)
    return str(v).strip()


def _parse_date_from_cell(v: Any) -> Optional[date]:
    """Parse date from cell (could be datetime or string). Returns date or None."""
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if hasattr(v, "date") and callable(getattr(v, "date")):
        d = v.date()
        return d if isinstance(d, date) else None
    return _parse_date(v)


def _load_xlsx_sheet(path: str, sheet_name: str) -> Optional[Tuple[List[str], List[List]]]:
    """Load one sheet from .xlsx. Returns (headers, rows) or None."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            return None
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return None
        headers = [_normalize_header(str(h)) for h in rows[0]]
        data_rows = rows[1:]
        return (headers, data_rows)
    except Exception as e:
        logger.warning("openpyxl load %s sheet %s: %s", path, sheet_name, e)
        return None


def _load_xls_sheet(path: str, sheet_name: str) -> Optional[Tuple[List[str], List[List]]]:
    """Load one sheet from .xls. Returns (headers, rows) or None."""
    try:
        import xlrd
        wb = xlrd.open_workbook(path)
        sheet = None
        for i in range(wb.nsheets):
            if wb.sheet_by_index(i).name == sheet_name:
                sheet = wb.sheet_by_index(i)
                break
        if sheet is None:
            return None
        rows = []
        for r in range(sheet.nrows):
            row = []
            for c in range(sheet.ncols):
                cell = sheet.cell_value(r, c)
                if isinstance(cell, float) and cell == int(cell):
                    cell = int(cell)
                row.append(cell)
            rows.append(row)
        if not rows:
            return None
        headers = [_normalize_header(str(h)) for h in rows[0]]
        data_rows = [list(r) for r in rows[1:]]
        return (headers, data_rows)
    except Exception as e:
        logger.warning("xlrd load %s sheet %s: %s", path, sheet_name, e)
        return None


def _find_col(headers: List[str], name: str) -> Optional[int]:
    """Find first column index by normalized header name. Returns None if not found."""
    for i, h in enumerate(headers):
        if h == name or (name in h and h.startswith(name)):
            return i
    return None


def _find_col_infokom_personal_number(headers: List[str]) -> Optional[int]:
    """On Infokom, 'табельный №' appears twice; we need the FIRST one (personal number)."""
    for i, h in enumerate(headers):
        if h == "табельный №":
            return i
    return None


def _find_col_infokom_full_name(headers: List[str]) -> Optional[int]:
    """On Infokom, 'табельный номер' column contains FIO (full name)."""
    return _find_col(headers, "табельный номер")


def parse_ddj(path: str) -> Tuple[List[Dict], List[str]]:
    """Parse sheet ДДЖ. Returns (list of row dicts keyed by field name, list of errors)."""
    out: List[Dict] = []
    errors: List[str] = []
    if path.lower().endswith(".xlsx"):
        data = _load_xlsx_sheet(path, SHEET_DDJ)
    else:
        data = _load_xls_sheet(path, SHEET_DDJ)
    if not data:
        return ([], [f"Sheet '{SHEET_DDJ}' not found or empty"])
    headers, data_rows = data
    col_map = {}
    for db_field, possible_headers in [
        ("personal_number", ["табельный номер"]),
        ("full_name", ["фои", "фио"]),
        ("position", ["должность"]),
        ("mvz", ["текст основное мвз"]),
        ("hire_date", ["дата приема"]),
        ("email", ["почта"]),
        ("supervisor", ["руководитель"]),
    ]:
        for h in possible_headers:
            idx = _find_col(headers, h)
            if idx is not None:
                col_map[db_field] = idx
                break
    if "personal_number" not in col_map or "full_name" not in col_map:
        return ([], [f"Sheet ДДЖ: required columns (Табельный номер, ФИО) not found"])
    for r_idx, row in enumerate(data_rows):
        personal_number = _cell_value(row, col_map.get("personal_number"))
        full_name = _cell_value(row, col_map.get("full_name"))
        if not personal_number and not full_name:
            continue
        if not personal_number:
            errors.append(f"ДДЖ row {r_idx + 2}: missing personal number")
            continue
        hire_val = None
        if col_map.get("hire_date") is not None and col_map["hire_date"] < len(row):
            hire_val = row[col_map["hire_date"]]
        rec = {
            "personal_number": personal_number,
            "full_name": full_name or "",
            "position": _cell_value(row, col_map.get("position")),
            "mvz": _cell_value(row, col_map.get("mvz")),
            "hire_date": _parse_date_from_cell(hire_val),
            "email": _cell_value(row, col_map.get("email")),
            "supervisor": _cell_value(row, col_map.get("supervisor")),
        }
        out.append(rec)
    return (out, errors)


def parse_infokom(path: str) -> Tuple[List[Dict], List[str]]:
    """Parse sheet Инфоком. Column 'Табельный номер' = full_name (FIO). First 'Табельный №' = personal_number."""
    out: List[Dict] = []
    errors: List[str] = []
    if path.lower().endswith(".xlsx"):
        data = _load_xlsx_sheet(path, SHEET_INFOKOM)
    else:
        data = _load_xls_sheet(path, SHEET_INFOKOM)
    if not data:
        return ([], [f"Sheet '{SHEET_INFOKOM}' not found or empty"])
    headers, data_rows = data
    col_personal = _find_col_infokom_personal_number(headers)
    col_full_name = _find_col_infokom_full_name(headers)
    col_position = _find_col(headers, INFOKOM_POSITION_COL)
    col_hire = _find_col(headers, INFOKOM_HIRE)
    col_supervisor = _find_col(headers, INFOKOM_SUPERVISOR) or _find_col(headers, _normalize_header("ФИО руководителя"))
    col_mvz = _find_col(headers, INFOKOM_MVZ)
    col_email = _find_col(headers, INFOKOM_EMAIL)
    if col_personal is None or col_full_name is None:
        return ([], [f"Sheet Инфоком: required columns (Табельный №, Табельный номер) not found"])
    for r_idx, row in enumerate(data_rows):
        personal_number = _cell_value(row, col_personal)
        full_name = _cell_value(row, col_full_name)  # on Infokom this column is FIO
        if not personal_number and not full_name:
            continue
        if not personal_number:
            errors.append(f"Инфоком row {r_idx + 2}: missing personal number")
            continue
        rec = {
            "personal_number": personal_number,
            "full_name": full_name or "",
            "position": _cell_value(row, col_position) if col_position is not None else "",
            "mvz": _cell_value(row, col_mvz) if col_mvz is not None else "",
            "hire_date": _parse_date_from_cell(row[col_hire] if col_hire is not None and col_hire < len(row) else None),
            "email": _cell_value(row, col_email) if col_email is not None else "",
            "supervisor": _cell_value(row, col_supervisor) if col_supervisor is not None else "",
        }
        out.append(rec)
    return (out, errors)


def _merge_rows_by_personal(ddj_rows: List[Dict], infokom_rows: List[Dict]) -> Dict[str, Dict]:
    """Merge two lists by personal_number. One number -> one record; fill blanks from the other sheet."""
    by_num: Dict[str, Dict] = {}
    for r in ddj_rows:
        pn = (r.get("personal_number") or "").strip()
        if pn:
            by_num[pn] = dict(r)
    for r in infokom_rows:
        pn = (r.get("personal_number") or "").strip()
        if not pn:
            continue
        if pn in by_num:
            existing = by_num[pn]
            for k, v in r.items():
                if v and (not existing.get(k) or existing.get(k) == ""):
                    existing[k] = v
        else:
            by_num[pn] = dict(r)
    return by_num


def import_employees_from_file(file_path: str) -> Any:
    """
    Parse .xlsx/.xls (sheets ДДЖ and Инфоком), check duplicate personal_number in file,
    merge by personal_number, insert only new employees. Optionally run Jira enrichment (Task 5).
    Returns dict: added_count, added_names, errors, [enrichment_errors]
    or "Error: ..." string on fatal error (e.g. duplicate in file, file not found).
    """
    path = Path(file_path).resolve()
    if not path.exists() or not path.is_file():
        return "Error: File not found."
    suf = path.suffix.lower()
    if suf not in (".xlsx", ".xls"):
        return "Error: Only .xlsx and .xls files are supported."

    ddj_rows, ddj_err = parse_ddj(str(path))
    infokom_rows, inf_err = parse_infokom(str(path))
    errors = ddj_err + inf_err

    # Build set of personal numbers per sheet to detect duplicates (same number on both sheets or twice on one)
    ddj_nums: Dict[str, List[str]] = {}
    for r in ddj_rows:
        pn = (r.get("personal_number") or "").strip()
        if pn:
            ddj_nums.setdefault(pn, []).append("ДДЖ")
    inf_nums: Dict[str, List[str]] = {}
    for r in infokom_rows:
        pn = (r.get("personal_number") or "").strip()
        if pn:
            inf_nums.setdefault(pn, []).append("Инфоком")
    # Duplicate = same number in both sheets, or same number more than once in one sheet
    for pn, locs in ddj_nums.items():
        if len(locs) > 1:
            return f"Error: Duplicate personal number in file: '{pn}' on sheet ДДЖ (rows with this number). Import aborted."
        if pn in inf_nums:
            return f"Error: Duplicate personal number in file: '{pn}' appears on both sheets (ДДЖ and Инфоком). Import aborted."
    for pn, locs in inf_nums.items():
        if len(locs) > 1:
            return f"Error: Duplicate personal number in file: '{pn}' on sheet Инфоком (rows with this number). Import aborted."

    merged = _merge_rows_by_personal(ddj_rows, infokom_rows)
    added: List[Dict] = []
    for pn, rec in merged.items():
        if not pn or not (rec.get("full_name") or rec.get("email")):
            continue
        if employee_exists_by_personal_number(pn):
            continue
        email = (rec.get("email") or "").strip()
        full_name = (rec.get("full_name") or "").strip()
        if not email:
            errors.append(f"Personal number {pn} ({full_name}): missing email, skip insert")
            continue
        try:
            created = create_employee(
                personal_number=pn,
                full_name=full_name,
                email=email,
                position=(rec.get("position") or "").strip() or None,
                mvz=(rec.get("mvz") or "").strip() or None,
                supervisor=(rec.get("supervisor") or "").strip() or None,
                hire_date=rec.get("hire_date"),
                mattermost_username=email,
            )
            added.append(created)
        except Exception as e:
            errors.append(f"Personal number {pn}: {e!s}")

    # Jira enrichment for newly added (will be implemented in Task 5)
    enrichment_errors: List[str] = []
    try:
        from plugins.hr_service.jira_enrichment import enrich_new_employees_jira
        enriched, enrichment_errors = enrich_new_employees_jira([a["id"] for a in added])
        logger.info("Jira enrichment: %d enriched, %d errors", enriched, len(enrichment_errors))
    except Exception as e:
        logger.warning("Jira enrichment not run: %s", e)

    added_names = [a.get("full_name") or a.get("personal_number", "") for a in added]
    result = {
        "added_count": len(added),
        "added_names": added_names[:10] if len(added_names) <= 10 else [],
        "errors": errors,
        "enrichment_errors": enrichment_errors,
    }
    return result
